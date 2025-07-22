import tkinter as tk
from tkinter import filedialog, messagebox
import os
import sys
import sqlite3
from constants import MEASUREMENTS, parse_ingredient
from openpyxl import load_workbook
import xlwt


        
def load_recipes_from_xls(file_path):
    import re  # Import regex for parsing numbers
    try:
        if not file_path:
            print("No file selected.")
            return

        # Connect to the database
        conn = sqlite3.connect('mealMaestro_data.db')
        cursor = conn.cursor()

        # Load the Excel file
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Fetch all ingredient names from ingredient_dict
        cursor.execute("SELECT ingredient_name FROM ingredient_dict")
        ingredient_dict = [row[0] for row in cursor.fetchall()]

        # Iterate through rows in the Excel file
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
            ingredient_description, quantity, measurement, recipe_name = row

            # Parse the ingredient description
            if isinstance(ingredient_description, str):
                ingredient_description = ingredient_description.strip()

                # Check if the ingredient starts with a number
                match = re.match(r'^(\d+\.?\d*)\s*([a-zA-Z]*)\s*(.*)', ingredient_description)
                if match:
                    quantity = match.group(1)  # Extract the quantity
                    measurement = match.group(2)  # Extract the measurement
                    ingredient_description = match.group(3)  # Extract the remaining description
                else:
                    # If no number is found, treat the entire value as the description
                    quantity = None
                    measurement = None

            # Determine ingredient_name by matching ingredient_description with ingredient_dict
            ingredient_name = None
            for name in ingredient_dict:
                if name.lower() in ingredient_description.lower():
                    ingredient_name = name
                    break

            # Insert recipe into the database (if it doesn't already exist)
            cursor.execute('SELECT id FROM recipes WHERE name = ?', (recipe_name,))
            recipe = cursor.fetchone()
            if not recipe:
                cursor.execute('INSERT INTO recipes (name) VALUES (?)', (recipe_name,))
                recipe_id = cursor.lastrowid
            else:
                recipe_id = recipe[0]

            # Insert ingredient into the database
            cursor.execute('''
                INSERT INTO ingredients (recipe_id, ingredient_description, quantity, measurement, ingredient_name)
                VALUES (?, ?, ?, ?, ?)
            ''', (recipe_id, ingredient_description, quantity, measurement, ingredient_name))

        conn.commit()
        print(f"Recipes and ingredients loaded successfully from {file_path}.")
    except Exception as e:
        print(f"Failed to load recipes from Excel: {e}")
    finally:
        conn.close()

def dump_recipes_to_xls(file_path):
    try:
        if not file_path:
            print("No file specified.")
            return

        # Connect to the database
        conn = sqlite3.connect('mealMaestro_data.db')
        cursor = conn.cursor()

        # Create a workbook and add a sheet
        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet("Recipes Dump")

        # Write headers
        sheet.write(0, 0, "Ingredient Description")
        sheet.write(0, 1, "Quantity")
        sheet.write(0, 2, "Measurement")
        sheet.write(0, 3, "Recipe Name")

        row = 1  # Start writing data from row 1

        # Fetch all recipes and their ingredients
        cursor.execute('''
            SELECT 
                ingredients.quantity AS quantity,
                ingredients.measurement AS measurement,
                ingredients.ingredient_description AS ingredient_description,
                recipes.name AS recipe_name
            FROM recipes
            LEFT JOIN ingredients ON recipes.id = ingredients.recipe_id
            ORDER BY recipes.name, ingredients.ingredient_description
        ''')
        rows = cursor.fetchall()

        # Write data to the Excel sheet
        for quantity, measurement, ingredient_description, recipe_name in rows:
            sheet.write(row, 0, ingredient_description)
            sheet.write(row, 1, quantity if quantity is not None else "")
            sheet.write(row, 2, measurement if measurement is not None else "")
            sheet.write(row, 3, recipe_name)
            row += 1

        # Save the workbook as an .xls file
        workbook.save(file_path)
        print(f"Recipes dumped successfully to {file_path}.")
    except Exception as e:
        print(f"Failed to dump recipes to Excel: {e}")
    finally:
        conn.close()