import sqlite3
from collections import defaultdict
from constants import MEASUREMENTS, parse_ingredient
import xlwt  # Import xlwt for writing Excel files
from recipeUtil import auto_resize_popup
from openpyxl import load_workbook  # Import load_workbook to read Excel files
from fractions import Fraction  # Import Fraction to handle fractional quantities
from openpyxl import Workbook  # Import Workbook to create Excel files
import tkinter as tk
from tkinter import filedialog, messagebox

# Function to create a meal plan
def create_meal_plan(root):
    def save_meal_plan():
        title = title_entry.get()
        selected_recipes = recipe_listbox.curselection()
        if title and selected_recipes:
            recipe_ids = ', '.join([str(recipe_list[i][0]) for i in selected_recipes])  # Get selected recipe IDs
            try:
                conn = sqlite3.connect('mealMaestro_data.db')
                cursor = conn.cursor()
                cursor.execute('''
                    INSERT INTO meal_plans (title, recipe_ids)
                    VALUES (?, ?)
                ''', (title, recipe_ids))
                conn.commit()
                conn.close()
                messagebox.showinfo("Success", "Meal plan saved successfully!")
                create_meal_plan_window.destroy()
            except sqlite3.IntegrityError:
                messagebox.showerror("Error", "Meal plan title must be unique.")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save meal plan: {e}")
        else:
            messagebox.showerror("Error", "Please enter a title and select at least one recipe.")

    def load_recipes():
        try:
            conn = sqlite3.connect('mealMaestro_data.db')
            cursor = conn.cursor()
            cursor.execute('SELECT id, name FROM recipes')
            global recipe_list
            recipe_list = cursor.fetchall()
            conn.close()
            for recipe in recipe_list:
                recipe_listbox.insert(tk.END, recipe[1])  # Add recipe names to the listbox
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load recipes: {e}")

    create_meal_plan_window = tk.Toplevel(root)
    create_meal_plan_window.title("Create Meal Plan")
    create_meal_plan_window.geometry("400x400")

    tk.Label(create_meal_plan_window, text="Meal Plan Title or Date:").pack(pady=5)
    title_entry = tk.Entry(create_meal_plan_window, width=30)
    title_entry.pack(pady=5)

    tk.Label(create_meal_plan_window, text="Select Recipes:").pack(pady=5)
    recipe_listbox = tk.Listbox(create_meal_plan_window, width=40, height=10, selectmode=tk.MULTIPLE)  # Allow multiple selection
    recipe_listbox.pack(pady=5)

    load_recipes()  # Load recipes into the listbox

    tk.Button(create_meal_plan_window, text="Save Meal Plan", command=save_meal_plan).pack(pady=10)

def generate_shopping_list(root):
    # Create a popup window for selecting a meal plan
    shopping_list_window = tk.Toplevel(root)
    shopping_list_window.title("Generate Shopping List")
    shopping_list_window.geometry("400x200")

    # Fetch meal plans from the database
    conn = sqlite3.connect('mealMaestro_data.db')
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT id, title FROM meal_plans")
        meal_plans = cursor.fetchall()  # List of (id, title)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to fetch meal plans: {e}")
        return
    finally:
        conn.close()

    # If no meal plans are available, show an error
    if not meal_plans:
        messagebox.showerror("Error", "No meal plans available.")
        return

    # Dropdown to select a meal plan
    tk.Label(shopping_list_window, text="Select a Meal Plan:").pack(pady=10)
    selected_meal_plan = tk.StringVar(shopping_list_window)
    selected_meal_plan.set(meal_plans[0][1])  # Set the first meal plan as default

    dropdown = tk.OptionMenu(shopping_list_window, selected_meal_plan, *[plan[1] for plan in meal_plans])
    dropdown.pack(pady=5)

    def save_shopping_list():
        meal_plan_title = selected_meal_plan.get()

        # Fetch recipes associated with the selected meal plan
        conn = sqlite3.connect('mealMaestro_data.db')
        cursor = conn.cursor()
        try:
            cursor.execute('''
                SELECT recipe_ids
                FROM meal_plans
                WHERE title = ?
            ''', (meal_plan_title,))
            result = cursor.fetchone()
            if result:
                recipe_ids = result[0].split(",")  # Recipe IDs are stored as a comma-separated string
            else:
                recipe_ids = []
        except Exception as e:
            messagebox.showerror("Error", f"Failed to fetch recipes for the meal plan: {e}")
            return
        finally:
            conn.close()

        # Fetch ingredients for the selected recipes, including recipe names
        conn = sqlite3.connect('mealMaestro_data.db')
        cursor = conn.cursor()
        try:
            ingredients = []
            for recipe_id in recipe_ids:
                cursor.execute('''
                    SELECT r.name AS recipe_name, 
                        i.ingredient_description AS ingredient, 
                        SUM(CAST(i.quantity AS REAL)) AS total_quantity, 
                        i.measurement
                    FROM ingredients i
                    JOIN recipes r ON i.recipe_id = r.id
                    WHERE i.recipe_id = ?
                    GROUP BY r.name, i.ingredient_description, i.measurement
                ''', (recipe_id,))
                ingredients.extend(cursor.fetchall())
        except Exception as e:
            messagebox.showerror("Error", f"Failed to fetch ingredients: {e}")
            return
        finally:
            conn.close()

        # # Prompt the user to save the shopping list as an Excel file
        # file_path = filedialog.asksaveasfilename(
        #     title="Save Shopping List",
        #     defaultextension=".xlsx",
        #     filetypes=[("Excel Files", "*.xlsx")]
        # )
        # if not file_path:
        #     messagebox.showerror("Error", "No file selected for saving.")
        #     return

        # Load the ingredient dictionary from ingredient_dict.xlsx
        ingredient_dict = {}
        try:
            ingredient_workbook = load_workbook("ingredient_dict.xlsx")
            ingredient_sheet = ingredient_workbook.active
            for row in ingredient_sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
                if row[0] and row[1]:  # Ensure both key and value exist
                    ingredient_dict[row[0].strip().lower()] = row[1].strip()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load ingredient dictionary: {e}")
            return

        # Load the store dictionary from store_lookup.xlsx
        store_dict = {}
        try:
            store_workbook = load_workbook("store_lookup.xlsx")
            store_sheet = store_workbook.active
            for row in store_sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
                if row[0] and row[1]:  # Ensure both key and value exist
                    store_dict[row[0].strip().lower()] = row[1].strip()
            print("DEBUG: Store dictionary loaded successfully.")
        except Exception as e:
            print(f"DEBUG: Failed to load store dictionary: {e}")
            messagebox.showerror("Error", f"Failed to load store dictionary: {e}")
            return

        # Create the Excel file
        try:
            from fractions import Fraction  # Import Fraction to handle fractional quantities

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Shopping List"

 
            # Write headers
            sheet.append(["Recipe Name", "Ingredient", "Total Quantity", "Measurement", "Generic Ingredient", "Store"])

            # Load the store_lookup.xlsx file and use it for both the generic ingredient dictionary and store lookup
            generic_ingredient_dict = {}  # Dictionary to map ingredient descriptions to generic ingredients
            store_dict = {}  # Dictionary to map generic ingredients to stores
            try:
                store_workbook = load_workbook("store_lookup.xlsx")
                store_sheet = store_workbook.active
                for row in store_sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
                    if row[0] and row[1]:  # Ensure both columns have values
                        generic_ingredient = row[0].strip().lower()
                        store = row[1].strip()
                        generic_ingredient_dict[generic_ingredient] = generic_ingredient  # Map generic ingredient to itself
                        store_dict[generic_ingredient] = store  # Map generic ingredient to store
                print("DEBUG: Loaded store_lookup.xlsx successfully.")
            except Exception as e:
                print(f"DEBUG: Failed to load store_lookup.xlsx: {e}")
                messagebox.showerror("Error", f"Failed to load store_lookup.xlsx: {e}")
                return

            # Write ingredients to the Excel file
            for recipe_name, ingredient, quantity, measurement in ingredients:
                try:
                    if quantity is None:  # Handle None values
                        total_quantity = ""
                    else:
                        # Ensure quantity is treated as a string
                        quantity_str = str(quantity).strip()

                        if "to" in quantity_str:  # Handle ranges like "1/4 to 1/2" or "1/4 to 2"
                            parts = quantity_str.split("to")
                            total_quantity = sum(float(Fraction(part.strip())) for part in parts) / len(parts)
                        elif " " in quantity_str:  # Handle mixed fractions like "1 1/2"
                            parts = quantity_str.split(" ")
                            total_quantity = float(parts[0]) + float(Fraction(parts[1]))
                        else:  # Handle single fractions like "1/2"
                            total_quantity = float(Fraction(quantity_str))
                except (ValueError, ZeroDivisionError):
                    total_quantity = quantity  # Keep as-is if parsing fails

                # Find the generic ingredient from the dictionary
                generic_ingredient = ""
                for key in generic_ingredient_dict.keys():
                    #print(f"DEBUG: Checking if '{key}' is in ingredient '{ingredient.lower()}'.")  # Debug output for each dictionary key
                    if key in ingredient.lower():  # Check if the dictionary key exists in the ingredient description
                        #print(f"DEBUG: Found match for '{key}' in ingredient '{ingredient}'. Generic ingredient: '{key}'")
                        generic_ingredient = key
                        break
                else:
                    print(f"DEBUG: No match found for ingredient '{ingredient}' in the dictionary.")  # Debug output for no match

                # Lookup the store for the generic ingredient
                store = ""
                if generic_ingredient:
                    store = store_dict.get(generic_ingredient, "")  # Lookup the store in the store dictionary
                    if store:
                        print(f"DEBUG: Found store '{store}' for generic ingredient '{generic_ingredient}'")
                    else:
                        print(f"DEBUG: No store found for generic ingredient '{generic_ingredient}'")

                # Append the processed ingredient to the Excel sheet
                sheet.append([recipe_name, ingredient, total_quantity or "", measurement or "", generic_ingredient, store])

            # Save the Excel file
            file_path = "shopping_list.xlsx"
            try:
                workbook.save(file_path)
                messagebox.showinfo("Success", f"Shopping list saved to {file_path}.")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save shopping list: {e}")
            shopping_list_window.destroy()

        except Exception as e:
            messagebox.showerror("Error", f"Failed to save shopping list: {e}")

    # Submit button to generate and save the shopping list
    submit_button = tk.Button(shopping_list_window, text="Generate Shopping List", command=save_shopping_list)
    submit_button.pack(pady=10)