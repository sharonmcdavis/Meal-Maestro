import sqlite3
from openpyxl import load_workbook

def setup_database():
    conn = sqlite3.connect('mealMaestro_data.db')
    cursor = conn.cursor()
    
    # Create recipes table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS recipes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS ingredients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            recipe_id INTEGER NOT NULL,
            quantity TEXT,
            measurement TEXT,
            ingredient_description TEXT NOT NULL,
            FOREIGN KEY (recipe_id) REFERENCES recipes (id)
        )
    ''')
    
    # Create meal_plans table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS meal_plans (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL UNIQUE,
            recipe_ids TEXT NOT NULL,
            FOREIGN KEY (id) REFERENCES recipes (id)
        )
    ''')
    
    # Create shopping_lists table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS shopping_lists (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ingredient TEXT NOT NULL
        )
    ''')
    
        # Create ingredient_dict table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS ingredient_dict (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ingredient_name TEXT NOT NULL UNIQUE
        )
    ''')

    conn.commit()
    conn.close()
    print("Database setup complete.")


def import_ingredient_dict():
    try:
        # Automatically set the file path to ingredient_dict.xls in the current directory
        file_path = "ingredient_dict.xls"

        # Check if the file exists
        import os
        if not os.path.exists(file_path):
            print(f"File {file_path} not found.")
            return

        # Connect to the database
        conn = sqlite3.connect('mealMaestro_data.db')
        cursor = conn.cursor()

        # Load the Excel file
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Iterate through rows in the Excel file
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
            ingredient_name = row[0]  # Assuming the first column contains ingredient names

            # Insert ingredient into the ingredient_dict table
            cursor.execute('''
                INSERT OR IGNORE INTO ingredient_dict (ingredient_name)
                VALUES (?)
            ''', (ingredient_name,))

        conn.commit()
        print(f"Ingredient dictionary imported successfully from {file_path}.")
    except Exception as e:
        print(f"Failed to import ingredient dictionary: {e}")
    finally:
        conn.close()

if __name__ == "__main__":
    setup_database()
    import_ingredient_dict()